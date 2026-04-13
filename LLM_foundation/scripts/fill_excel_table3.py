import pandas as pd
from openpyxl import load_workbook
import os

xlsx_path = r'c:\Projectbulid\LLM_foundation\paper1_output\results\Publication_Tables_Final.xlsx'
csv_path = r'c:\Projectbulid\LLM_foundation\paper1_output\results\Table_DVS_final_unified.csv'

# Load the CSV
unified_df = pd.read_csv(csv_path)

# Load the workbook
wb = load_workbook(xlsx_path)
ws = wb['Table 3 DVS']

# Find header column index for 'Miss if Yes' and 'Miss if No' and 'Variable'
# We'll just look at the first few rows to find where they are
header_row = 1
for r in range(1, 10):
    row_vals = [str(cell.value).strip() if cell.value else '' for cell in ws[r]]
    if any('Variable' in v for v in row_vals):
        header_row = r
        break

header_vals = [str(cell.value).strip() if cell.value else '' for cell in ws[header_row]]
try:
    var_col = next(i for i, v in enumerate(header_vals) if 'Variable' in v)
    yes_col = next(i for i, v in enumerate(header_vals) if 'Miss if Yes' in v or 'present' in v.lower())
    no_col = next(i for i, v in enumerate(header_vals) if 'Miss if No' in v or 'absent' in v.lower())
except StopIteration:
    # If explicit headers aren't perfectly named, we fall back to indices based on common structure
    # Usually Variable is col 1, Miss if Yes is col 3, Miss if No is col 4
    # Let's just find them by index and assign safely
    var_col = 0
    yes_col = 2
    no_col = 3

# Inject values
for r in range(header_row + 1, ws.max_row + 1):
    var_name = str(ws.cell(row=r, column=var_col+1).value).strip()
    if var_name:
        # Match with CSV
        match = unified_df[unified_df['Variable'].astype(str).str.contains(var_name, case=False, na=False, regex=False)]
        idx = unified_df['Variable'].apply(lambda x: str(x).strip() in var_name or var_name in str(x).strip())
        match = unified_df[idx]
        
        if not match.empty:
            ws.cell(row=r, column=yes_col+1).value = match.iloc[0]['Miss if Yes']
            ws.cell(row=r, column=no_col+1).value = match.iloc[0]['Miss if No']

wb.save(xlsx_path)
print("Table 3 successfully populated.")

